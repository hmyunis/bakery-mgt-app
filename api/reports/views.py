from datetime import datetime
from decimal import Decimal

import openpyxl
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncHour
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from production.models import IngredientUsage, Product, ProductionRun

# --- App Imports ---
from sales.models import Sale, SaleItem, SalePayment
from users.models import (
    AttendanceRecord,
    Employee,
    LeaveRecord,
    PayrollRecord,
    ShiftAssignment,
)

# ==========================================
# HELPER: Excel Styling & Formatting
# ==========================================


class ExcelStyler:
    """
    Handles design: Big titles, readable tables, simple colors.
    """

    def __init__(self):
        # Professional but friendly colors (Navy Blue & White theme)
        self.header_bg = "203764"  # Deep Navy Blue
        self.header_text = "FFFFFF"  # White
        self.total_bg = "DDEBF7"  # Very Light Blue
        self.zebra_stripe = "F7F7F7"  # Almost white grey

        # Borders
        self.thin_border = Side(border_style="thin", color="BFBFBF")
        self.thick_border = Side(border_style="medium", color="000000")
        self.border_full = Border(
            left=self.thin_border,
            right=self.thin_border,
            top=self.thin_border,
            bottom=self.thin_border,
        )

        # Fonts
        self.font_title = Font(name="Calibri", size=18, bold=True, color="203764")
        self.font_subtitle = Font(name="Calibri", size=12, italic=True, color="595959")
        self.font_header = Font(
            name="Calibri", size=11, bold=True, color=self.header_text
        )
        self.font_total = Font(name="Calibri", size=11, bold=True)
        self.font_body = Font(name="Calibri", size=11)

        # Alignments
        self.align_center = Alignment(horizontal="center", vertical="center")
        self.align_left = Alignment(horizontal="left", vertical="center")
        self.align_right = Alignment(horizontal="right", vertical="center")

    def add_sheet_header(self, ws, title, subtitle, merge_cols=6):
        """Adds the big text at the top of the sheet"""
        ws["A1"] = title
        ws["A1"].font = self.font_title

        ws["A2"] = subtitle
        ws["A2"].font = self.font_subtitle

        # Merge cells slightly so text doesn't get cut off
        end_col = get_column_letter(max(1, merge_cols))
        ws.merge_cells(f"A1:{end_col}1")
        ws.merge_cells(f"A2:{end_col}2")

    def create_table(self, ws, headers, data, sum_columns=None):
        """
        Standard table generator.
        """
        if sum_columns is None:
            sum_columns = []

        start_row = 4

        # 1. Write Headers
        ws.append([])
        ws.append(headers)  # Row 4

        for cell in ws[start_row]:
            cell.fill = PatternFill(
                start_color=self.header_bg, end_color=self.header_bg, fill_type="solid"
            )
            cell.font = self.font_header
            cell.alignment = self.align_center
            cell.border = self.border_full

        # 2. Write Data
        totals = {col_idx: 0.0 for col_idx in sum_columns}

        for i, row_data in enumerate(data):
            ws.append(row_data)
            current_row = ws[ws.max_row]

            # Zebra Striping
            is_even = i % 2 == 0
            fill = (
                PatternFill(
                    start_color=self.zebra_stripe,
                    end_color=self.zebra_stripe,
                    fill_type="solid",
                )
                if is_even
                else None
            )

            for col_idx, cell in enumerate(current_row):
                cell.font = self.font_body
                cell.border = self.border_full
                if fill:
                    cell.fill = fill

                if isinstance(cell.value, bool):
                    pass
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"
                elif isinstance(cell.value, (float, Decimal)):
                    cell.number_format = "#,##0.00"

                if col_idx in sum_columns and isinstance(
                    cell.value, (int, float, Decimal)
                ):
                    totals[col_idx] += float(cell.value)

        # 3. Write Totals Row
        if sum_columns and data:
            total_row_data = [""] * len(headers)
            total_row_data[0] = "GRAND TOTAL"

            for col_idx, total_val in totals.items():
                total_row_data[col_idx] = total_val

            ws.append(total_row_data)
            row_cells = ws[ws.max_row]

            for cell in row_cells:
                cell.font = self.font_total
                cell.fill = PatternFill(
                    start_color=self.total_bg,
                    end_color=self.total_bg,
                    fill_type="solid",
                )
                cell.border = Border(
                    top=self.thick_border,
                    bottom=self.thick_border,
                    left=self.thin_border,
                    right=self.thin_border,
                )
                if isinstance(cell.value, bool):
                    pass
                elif isinstance(cell.value, int):
                    cell.number_format = "#,##0"
                elif isinstance(cell.value, (float, Decimal)):
                    cell.number_format = "#,##0.00"

        self.adjust_column_widths(ws)

    def adjust_column_widths(self, ws):
        """
        Smart column width adjustment that handles Merged Cells correctly.
        """
        for col in ws.columns:
            max_length = 0
            # FIX: Use get_column_letter to avoid AttributeError on MergedCells
            column_letter = get_column_letter(col[0].column)

            for cell in col:
                # Skip title rows
                if cell.row < 3:
                    continue
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except Exception:
                    pass

            adjusted_width = max_length + 3
            if adjusted_width < 12:
                adjusted_width = 12
            if adjusted_width > 50:
                adjusted_width = 50

            ws.column_dimensions[column_letter].width = adjusted_width


# ==========================================
# VIEW 1: JSON Dashboard Stats
# ==========================================


class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_str = request.query_params.get("date")
        if not date_str:
            date_str = timezone.now().date().isoformat()

        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 1. Hourly Sales
        hourly_sales = (
            Sale.objects.filter(created_at__date=target_date)
            .annotate(hour=TruncHour("created_at"))
            .values("hour")
            .annotate(total=Sum("total_amount"))
            .order_by("hour")
        )

        hourly_data = []
        for item in hourly_sales:
            hourly_data.append(
                {"hour": item["hour"].strftime("%H:%M"), "total": float(item["total"])}
            )

        # 2. Top 5 Selling Products
        top_products = (
            SaleItem.objects.filter(sale__created_at__date=target_date)
            .values("product__name")
            .annotate(quantity=Sum("quantity"))
            .order_by("-quantity")[:5]
        )

        top_products_data = [
            {"name": item["product__name"], "quantity": float(item["quantity"])}
            for item in top_products
        ]

        # 3. Payment Method Distribution
        payment_methods = (
            SalePayment.objects.filter(sale__created_at__date=target_date)
            .values("method__name")
            .annotate(total=Sum("amount"))
            .order_by("-total")
        )

        payment_data = [
            {"name": item["method__name"], "value": float(item["total"])}
            for item in payment_methods
        ]

        # 4. Production vs Sales
        top_produced = (
            ProductionRun.objects.filter(
                date_produced__date=target_date, product__isnull=False
            )
            .values("product__name")
            .annotate(produced=Sum("quantity_produced"))
            .order_by("-produced")[:5]
        )

        prod_vs_sales_data = []
        for item in top_produced:
            product_name = item["product__name"]
            produced_qty = float(item["produced"])
            sold_qty = (
                SaleItem.objects.filter(
                    sale__created_at__date=target_date, product__name=product_name
                ).aggregate(total=Sum("quantity"))["total"]
                or 0
            )

            prod_vs_sales_data.append(
                {
                    "name": product_name,
                    "produced": produced_qty,
                    "sold": float(sold_qty),
                }
            )

        # 5. Wastage
        wastage = (
            IngredientUsage.objects.filter(
                production_run__date_produced__date=target_date
            )
            .values("ingredient__name")
            .annotate(total_wastage=Sum("wastage"))
            .order_by("-total_wastage")[:5]
        )

        wastage_data = [
            {"name": item["ingredient__name"], "wastage": float(item["total_wastage"])}
            for item in wastage
        ]

        # 6. Cashier
        cashier_perf = (
            Sale.objects.filter(created_at__date=target_date)
            .values("cashier__username", "cashier__full_name")
            .annotate(total_sales=Sum("total_amount"), count=Count("id"))
            .order_by("-total_sales")
        )

        cashier_data = [
            {
                "name": item["cashier__full_name"] or item["cashier__username"],
                "sales": float(item["total_sales"]),
                "count": item["count"],
            }
            for item in cashier_perf
        ]

        # 7. Stock
        products_in_stock = Product.objects.filter(stock_quantity__gt=0).count()

        return Response(
            {
                "hourly_sales": hourly_data,
                "top_products": top_products_data,
                "payment_methods": payment_data,
                "production_vs_sales": prod_vs_sales_data,
                "wastage": wastage_data,
                "cashier_performance": cashier_data,
                "products_in_stock": products_in_stock,
            }
        )


# ==========================================
# VIEW 2: Excel Report Export
# ==========================================


class ExportReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date_str = request.query_params.get("start_date")
        end_date_str = request.query_params.get("end_date")

        if not start_date_str or not end_date_str:
            return Response(
                {"error": "start_date and end_date are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response(
                {"error": "Invalid date format. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        wb = openpyxl.Workbook()
        styler = ExcelStyler()

        # ==========================================
        # SHEET 1: OVERVIEW
        # ==========================================
        ws_overview = wb.active
        ws_overview.title = "Business Snapshot"
        ws_overview.sheet_view.showGridLines = False

        total_money_in = (
            Sale.objects.filter(
                created_at__date__range=[start_date, end_date]
            ).aggregate(Sum("total_amount"))["total_amount__sum"]
            or 0
        )
        count_sales = Sale.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).count()
        avg_spend = total_money_in / count_sales if count_sales > 0 else 0

        try:
            from inventory.models import Purchase

            total_money_out = (
                Purchase.objects.filter(
                    purchase_date__date__range=[start_date, end_date]
                ).aggregate(Sum("total_cost"))["total_cost__sum"]
                or 0
            )
        except ImportError:
            total_money_out = 0

        total_payroll_paid = (
            PayrollRecord.objects.filter(
                paid_at__date__range=[start_date, end_date]
            ).aggregate(Sum("amount_paid"))["amount_paid__sum"]
            or 0
        )

        styler.add_sheet_header(
            ws_overview, "Business Snapshot", f"Report from {start_date} to {end_date}"
        )

        def add_summary_line(row, label, value, is_money=False, bold_value=False):
            cell_lbl = ws_overview.cell(row=row, column=2, value=label)
            cell_val = ws_overview.cell(row=row, column=3, value=value)

            cell_lbl.font = Font(name="Calibri", size=11)
            cell_lbl.border = Border(bottom=Side(style="thin", color="CCCCCC"))

            cell_val.font = Font(name="Calibri", size=11, bold=bold_value)
            cell_val.alignment = Alignment(horizontal="right")
            cell_val.border = Border(bottom=Side(style="thin", color="CCCCCC"))

            if is_money:
                cell_val.number_format = "#,##0.00"

        ws_overview["B5"] = "Money Coming In"
        ws_overview["B5"].font = Font(
            name="Calibri", bold=True, size=12, color="203764"
        )

        add_summary_line(6, "Total Sales Amount", float(total_money_in), True, True)
        add_summary_line(7, "Number of Sales Made", count_sales)
        add_summary_line(8, "Avg. Amount per Customer", float(avg_spend), True)

        ws_overview["B10"] = "Money Going Out"
        ws_overview["B10"].font = Font(
            name="Calibri", bold=True, size=12, color="203764"
        )

        add_summary_line(11, "Cost of Ingredients Bought", float(total_money_out), True)
        add_summary_line(12, "Payroll Paid (HR)", float(total_payroll_paid), True)

        ws_overview["B14"] = "Estimated Profit"
        ws_overview["B14"].font = Font(
            name="Calibri", bold=True, size=12, color="203764"
        )

        est_profit = (
            float(total_money_in) - float(total_money_out) - float(total_payroll_paid)
        )
        ws_overview["B15"] = "Sales minus Purchases minus Payroll"
        ws_overview["C15"] = est_profit
        ws_overview["C15"].number_format = "#,##0.00"
        ws_overview["C15"].font = Font(
            bold=True, size=12, color="006100" if est_profit >= 0 else "9C0006"
        )

        ws_overview.column_dimensions["A"].width = 5
        ws_overview.column_dimensions["B"].width = 30
        ws_overview.column_dimensions["C"].width = 25

        # ==========================================
        # SHEET 2: SALES LIST
        # ==========================================
        ws_sales = wb.create_sheet("Sales List")
        headers = [
            "Date",
            "Time",
            "Sale #",
            "Receipt Issued",
            "Cashier Name",
            "Items Sold",
            "Total Bill",
            "Amount Paid",
            "Change",
            "Payments",
        ]
        styler.add_sheet_header(
            ws_sales,
            "Detailed Sales List",
            "List of sales made in the selected period",
            merge_cols=len(headers),
        )

        sales_qs = (
            Sale.objects.filter(created_at__date__range=[start_date, end_date])
            .select_related("cashier")
            .prefetch_related("items__product", "payments__method")
            .order_by("-created_at")
        )

        data = []
        for s in sales_qs:
            items = ", ".join(
                [f"{i.product.name} ({i.quantity})" for i in s.items.all()]
            )
            cashier_name = (
                s.cashier.full_name
                if s.cashier and s.cashier.full_name
                else (s.cashier.username if s.cashier else "Unknown")
            )
            amount_paid = sum((p.amount for p in s.payments.all()), Decimal("0"))
            change = max(Decimal("0"), amount_paid - s.total_amount)
            payments_str = ", ".join(
                [
                    f"{p.method.name}: ETB {float(p.amount):.2f}"
                    for p in s.payments.all()
                ]
            )

            data.append(
                [
                    s.created_at.date(),
                    s.created_at.time().strftime("%H:%M"),
                    str(s.id),
                    "Yes" if s.receipt_issued else "No",
                    cashier_name,
                    items,
                    float(s.total_amount),
                    float(amount_paid),
                    float(change),
                    payments_str,
                ]
            )

        styler.create_table(ws_sales, headers, data, sum_columns=[6, 7, 8])

        # ==========================================
        # SHEET 3: BEST SELLERS
        # ==========================================
        ws_prod = wb.create_sheet("Best Sellers")

        headers = [
            "Item Name",
            "Count Sold",
            "Money Earned",
            "Count Made in Kitchen",
            "Current Stock",
        ]
        styler.add_sheet_header(
            ws_prod,
            "Product Performance",
            "Which items are making the most money?",
            merge_cols=len(headers),
        )

        products = Product.objects.filter(is_active=True)
        data = []

        for p in products:
            sold_agg = SaleItem.objects.filter(
                sale__created_at__date__range=[start_date, end_date], product=p
            ).aggregate(q=Sum("quantity"), r=Sum("subtotal"))

            prod_agg = ProductionRun.objects.filter(
                date_produced__date__range=[start_date, end_date], product=p
            ).aggregate(q=Sum("quantity_produced"))

            qty_sold = sold_agg["q"] or 0
            money_earned = float(sold_agg["r"] or 0)
            qty_made = prod_agg["q"] or 0

            if qty_sold > 0 or qty_made > 0:
                data.append(
                    [p.name, qty_sold, money_earned, qty_made, p.stock_quantity]
                )

        data.sort(key=lambda x: x[2], reverse=True)
        styler.create_table(ws_prod, headers, data, sum_columns=[1, 2, 3])

        # ==========================================
        # SHEET 4: KITCHEN ACTIVITY (Merged & Styled)
        # ==========================================
        ws_run = wb.create_sheet("Kitchen Activity")
        headers = [
            "Date & Time",
            "Chef Name",
            "Item Made",
            "Qty Made",
            "Ingredient Used",
            "Amount Used",
            "Amount Wasted",
        ]
        styler.add_sheet_header(
            ws_run,
            "Production & Waste",
            "What the kitchen made and what was wasted",
            merge_cols=len(headers),
        )

        # 1. Write Headers
        ws_run.append([])
        ws_run.append(headers)  # Row 4

        header_row_idx = 4
        for cell in ws_run[header_row_idx]:
            cell.fill = PatternFill(
                start_color=styler.header_bg,
                end_color=styler.header_bg,
                fill_type="solid",
            )
            cell.font = styler.font_header
            cell.alignment = styler.align_center
            cell.border = styler.border_full

        # 2. Fetch Data
        usages = (
            IngredientUsage.objects.filter(
                production_run__date_produced__date__range=[start_date, end_date]
            )
            .select_related("production_run", "ingredient")
            .order_by("-production_run__date_produced")
        )  # Sort by run is crucial for merging

        # 3. Custom Loop for Merging
        current_row = 5
        start_merge_row = 5
        last_run_id = None

        # Zebra Striping logic for groups requires tracking "group index"
        group_idx = 0
        fill_color = None

        total_made = 0
        total_used = 0
        total_wasted = 0

        # We append a 'dummy' None at the end to force the merge logic to
        # trigger for the last group
        usage_list = list(usages)
        usage_list.append(None)

        for u in usage_list:
            if u is None:
                # End of list, trigger last merge
                run_id = -1
            else:
                run_id = u.production_run.id

            # CHECK GROUP CHANGE
            if last_run_id is not None and run_id != last_run_id:
                # Merge the Previous Group (Cols A, B, C, D)
                if current_row > start_merge_row:
                    # Merge Date
                    ws_run.merge_cells(
                        start_row=start_merge_row,
                        start_column=1,
                        end_row=current_row - 1,
                        end_column=1,
                    )
                    # Merge Chef
                    ws_run.merge_cells(
                        start_row=start_merge_row,
                        start_column=2,
                        end_row=current_row - 1,
                        end_column=2,
                    )
                    # Merge Item
                    ws_run.merge_cells(
                        start_row=start_merge_row,
                        start_column=3,
                        end_row=current_row - 1,
                        end_column=3,
                    )
                    # Merge Qty
                    ws_run.merge_cells(
                        start_row=start_merge_row,
                        start_column=4,
                        end_row=current_row - 1,
                        end_column=4,
                    )

                    # Apply Vertical Center to the merged cells
                    ws_run.cell(row=start_merge_row, column=1).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws_run.cell(row=start_merge_row, column=2).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws_run.cell(row=start_merge_row, column=3).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws_run.cell(row=start_merge_row, column=4).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

                start_merge_row = current_row
                group_idx += 1

            if u is None:
                break  # Exit loop

            # Prepare Colors for this group
            if group_idx % 2 == 0:
                fill_color = PatternFill(
                    start_color=styler.zebra_stripe,
                    end_color=styler.zebra_stripe,
                    fill_type="solid",
                )
            else:
                fill_color = None

            # Get Values
            run = u.production_run
            item_name = (
                run.product.name
                if run.product
                else (
                    run.composite_ingredient.name
                    if run.composite_ingredient
                    else "Mix/Prep"
                )
            )
            chef_name = (
                run.chef.full_name
                if run.chef and run.chef.full_name
                else (run.chef.username if run.chef else "Unknown")
            )

            # Format: YYYY-MM-DD HH:MM
            dt_str = run.date_produced.strftime("%Y-%m-%d %H:%M")

            # Write Row
            ws_run.append(
                [
                    dt_str,
                    chef_name,
                    item_name,
                    float(run.quantity_produced),
                    u.ingredient.name,
                    float(u.actual_amount),
                    float(u.wastage),
                ]
            )

            # Apply Styles to the new row
            row_cells = ws_run[current_row]
            for col_idx, cell in enumerate(row_cells):
                cell.font = styler.font_body
                cell.border = styler.border_full
                if fill_color:
                    cell.fill = fill_color

                # Number formats
                if col_idx in [3, 5, 6]:  # Qty, Used, Wasted
                    cell.number_format = "#,##0.00"

            # Add to Totals (Only add Production Qty once per group)
            # Logic: If this is the FIRST row of a new group, add to production total.
            # Usage totals are always added.
            if run_id != last_run_id:
                total_made += run.quantity_produced

            total_used += u.actual_amount
            total_wasted += u.wastage

            last_run_id = run_id
            current_row += 1

        # 4. Totals Row for Kitchen
        ws_run.append(["GRAND TOTAL", "", "", total_made, "", total_used, total_wasted])

        last_row = ws_run[ws_run.max_row]
        for cell in last_row:
            cell.font = styler.font_total
            cell.fill = PatternFill(
                start_color=styler.total_bg,
                end_color=styler.total_bg,
                fill_type="solid",
            )
            cell.border = Border(
                top=styler.thick_border,
                bottom=styler.thick_border,
                left=styler.thin_border,
                right=styler.thin_border,
            )
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

        # Merge the empty cells in the total row for cleaner look (optional)
        ws_run.merge_cells(
            start_row=ws_run.max_row,
            start_column=1,
            end_row=ws_run.max_row,
            end_column=3,
        )

        styler.adjust_column_widths(ws_run)

        # ==========================================
        # SHEET 5: TEAM STATS
        # ==========================================
        ws_cashier = wb.create_sheet("Team Stats")

        headers = [
            "Staff Name",
            "Total Money Collected",
            "Customers Served",
            "Avg Sale Amount",
        ]
        styler.add_sheet_header(
            ws_cashier,
            "Staff Performance",
            "Who is selling the most?",
            merge_cols=len(headers),
        )

        c_stats = (
            Sale.objects.filter(created_at__date__range=[start_date, end_date])
            .values("cashier__full_name", "cashier__username")
            .annotate(total=Sum("total_amount"), count=Count("id"))
            .order_by("-total")
        )

        data = []
        for c in c_stats:
            name = c["cashier__full_name"] or c["cashier__username"] or "Unknown"
            tot = float(c["total"])
            cnt = c["count"]
            avg = tot / cnt if cnt else 0
            data.append([name, tot, cnt, avg])

        styler.create_table(ws_cashier, headers, data, sum_columns=[1, 2])

        # ==========================================
        # SHEET 6: HR SNAPSHOT
        # ==========================================
        ws_hr = wb.create_sheet("HR Snapshot")
        headers = ["Metric", "Value"]
        styler.add_sheet_header(
            ws_hr,
            "HR Snapshot",
            f"Staffing and payroll from {start_date} to {end_date}",
            merge_cols=len(headers),
        )

        employees_count = Employee.objects.count()
        total_monthly_salary = (
            Employee.objects.aggregate(Sum("monthly_base_salary"))[
                "monthly_base_salary__sum"
            ]
            or 0
        )

        scheduled_shifts = ShiftAssignment.objects.filter(
            shift_date__range=[start_date, end_date]
        ).count()

        attendance_qs = AttendanceRecord.objects.filter(
            assignment__shift_date__range=[start_date, end_date]
        )
        attendance_count = attendance_qs.count()
        attendance_by_status = {
            item["status"]: item["count"]
            for item in attendance_qs.values("status").annotate(count=Count("id"))
        }
        total_late_minutes = (
            attendance_qs.aggregate(Sum("late_minutes"))["late_minutes__sum"] or 0
        )
        total_overtime_minutes = (
            attendance_qs.aggregate(Sum("overtime_minutes"))["overtime_minutes__sum"]
            or 0
        )

        leaves_qs = LeaveRecord.objects.filter(
            Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
        ).select_related("employee")
        leave_count = leaves_qs.count()
        leave_days_in_period = 0
        for leave in leaves_qs:
            overlap_start = max(leave.start_date, start_date)
            overlap_end = min(leave.end_date, end_date)
            leave_days_in_period += (overlap_end - overlap_start).days + 1

        payroll_period_qs = PayrollRecord.objects.filter(
            Q(period_start__lte=end_date) & Q(period_end__gte=start_date)
        )
        payroll_period_count = payroll_period_qs.count()
        payroll_paid_count = payroll_period_qs.filter(
            status=PayrollRecord.STATUS_PAID
        ).count()
        payroll_unpaid_count = payroll_period_qs.filter(
            status=PayrollRecord.STATUS_UNPAID
        ).count()

        payroll_paid_in_period = (
            PayrollRecord.objects.filter(
                paid_at__date__range=[start_date, end_date]
            ).aggregate(Sum("amount_paid"))["amount_paid__sum"]
            or 0
        )

        hr_data = [
            ["Employees", employees_count],
            ["Total Monthly Base Salary", float(total_monthly_salary)],
            ["Shifts Scheduled", scheduled_shifts],
            ["Attendance Records", attendance_count],
            ["Present", attendance_by_status.get(AttendanceRecord.STATUS_PRESENT, 0)],
            ["Late", attendance_by_status.get(AttendanceRecord.STATUS_LATE, 0)],
            ["Absent", attendance_by_status.get(AttendanceRecord.STATUS_ABSENT, 0)],
            ["Overtime", attendance_by_status.get(AttendanceRecord.STATUS_OVERTIME, 0)],
            ["Total Late Minutes", total_late_minutes],
            ["Total Overtime Minutes", total_overtime_minutes],
            ["Leave Records", leave_count],
            ["Leave Days (in period)", leave_days_in_period],
            ["Payroll Records (overlapping)", payroll_period_count],
            ["Payroll Paid Records", payroll_paid_count],
            ["Payroll Unpaid Records", payroll_unpaid_count],
            ["Payroll Paid (in period)", float(payroll_paid_in_period)],
        ]
        styler.create_table(ws_hr, headers, hr_data)

        # ==========================================
        # SHEET 7: EMPLOYEES
        # ==========================================
        ws_emp = wb.create_sheet("Employees")
        headers = [
            "Employee ID",
            "Full Name",
            "Position",
            "Phone",
            "Hire Date",
            "Monthly Base Salary",
            "Payment Detail",
            "System Username",
            "System Role",
        ]
        styler.add_sheet_header(
            ws_emp, "Employees", "Employee directory", merge_cols=len(headers)
        )

        employees = Employee.objects.select_related("user").order_by("full_name")
        data = []
        for e in employees:
            data.append(
                [
                    e.id,
                    e.full_name,
                    e.position,
                    e.phone_number or "",
                    e.hire_date,
                    float(e.monthly_base_salary),
                    e.payment_detail or "",
                    e.user.username if e.user else "",
                    e.user.role if e.user else "",
                ]
            )
        styler.create_table(ws_emp, headers, data, sum_columns=[5])

        # ==========================================
        # SHEET 8: ATTENDANCE (Shift Attendance)
        # ==========================================
        ws_att = wb.create_sheet("Attendance")
        headers = [
            "Shift Date",
            "Employee",
            "Shift",
            "Start",
            "End",
            "Status",
            "Late (min)",
            "Overtime (min)",
            "Recorded By",
            "Notes",
        ]
        styler.add_sheet_header(
            ws_att,
            "Shift Attendance",
            f"Attendance records from {start_date} to {end_date}",
            merge_cols=len(headers),
        )

        assignments = (
            ShiftAssignment.objects.filter(shift_date__range=[start_date, end_date])
            .select_related("employee", "shift", "attendance__recorded_by")
            .order_by("shift_date", "shift__start_time", "employee__full_name")
        )

        data = []
        for a in assignments:
            att = getattr(a, "attendance", None)
            recorded_by = ""
            if att and att.recorded_by:
                recorded_by = att.recorded_by.full_name or att.recorded_by.username

            data.append(
                [
                    a.shift_date,
                    a.employee.full_name,
                    a.shift.name,
                    a.shift.start_time.strftime("%H:%M"),
                    a.shift.end_time.strftime("%H:%M"),
                    att.get_status_display() if att else "Not recorded",
                    att.late_minutes if att else 0,
                    att.overtime_minutes if att else 0,
                    recorded_by,
                    att.notes if att and att.notes else "",
                ]
            )

        styler.create_table(ws_att, headers, data, sum_columns=[6, 7])

        # ==========================================
        # SHEET 9: LEAVES
        # ==========================================
        ws_leave = wb.create_sheet("Leaves")
        headers = [
            "Employee",
            "Type",
            "Start Date",
            "End Date",
            "Days (Total)",
            "Days (in period)",
            "Notes",
        ]
        styler.add_sheet_header(
            ws_leave,
            "Leave Records",
            f"Leave records overlapping {start_date} to {end_date}",
            merge_cols=len(headers),
        )

        leaves = (
            LeaveRecord.objects.filter(
                Q(start_date__lte=end_date) & Q(end_date__gte=start_date)
            )
            .select_related("employee")
            .order_by("start_date", "employee__full_name")
        )

        data = []
        for leave in leaves:
            overlap_start = max(leave.start_date, start_date)
            overlap_end = min(leave.end_date, end_date)
            days_in_period = (overlap_end - overlap_start).days + 1
            data.append(
                [
                    leave.employee.full_name,
                    leave.get_leave_type_display(),
                    leave.start_date,
                    leave.end_date,
                    leave.day_count,
                    days_in_period,
                    leave.notes or "",
                ]
            )

        styler.create_table(ws_leave, headers, data, sum_columns=[4, 5])

        # ==========================================
        # SHEET 10: PAYROLL
        # ==========================================
        ws_pay = wb.create_sheet("Payroll")
        headers = [
            "Employee",
            "Period Start",
            "Period End",
            "Base Salary",
            "Amount Paid",
            "Outstanding",
            "Status",
            "Paid At",
            "Receipt Uploaded",
            "Notes",
        ]
        styler.add_sheet_header(
            ws_pay,
            "Payroll Records",
            f"Payroll records overlapping {start_date} to {end_date}",
            merge_cols=len(headers),
        )

        payrolls = (
            PayrollRecord.objects.filter(
                Q(period_start__lte=end_date) & Q(period_end__gte=start_date)
            )
            .select_related("employee")
            .order_by("period_start", "employee__full_name")
        )

        data = []
        for p in payrolls:
            outstanding = (p.base_salary or 0) - (p.amount_paid or 0)
            data.append(
                [
                    p.employee.full_name,
                    p.period_start,
                    p.period_end,
                    float(p.base_salary),
                    float(p.amount_paid),
                    float(outstanding),
                    p.get_status_display(),
                    p.paid_at.strftime("%Y-%m-%d %H:%M") if p.paid_at else "",
                    "Yes" if p.receipt else "No",
                    p.notes or "",
                ]
            )

        styler.create_table(ws_pay, headers, data, sum_columns=[3, 4, 5])

        # ==========================================
        # RETURN FILE
        # ==========================================
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"Shop_Report_{start_date}_{end_date}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        wb.save(response)
        return response
